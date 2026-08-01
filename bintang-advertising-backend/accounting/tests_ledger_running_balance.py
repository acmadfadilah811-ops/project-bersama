"""Rincian Mutasi Akun — saldo berjalan (running_balance) dan label Pelanggan/Supplier.

Bug ditemukan lewat contoh data referensi Olsera dari user: akun biaya (debit-normal,
mis. Pembelian) harus AKUMULASI dengan debit (saldo naik tiap ada debit), bukan
sebaliknya. Sebelumnya frontend (`RincianMutasiAkun.jsx`/`MutasiTable.jsx`)
menghitung ulang arah saldo sendiri lewat `isDebitNormal` yang dicek dari
`account.classification` — padahal endpoint `/accounting/ledger/<id>/` TIDAK
PERNAH mengirim `classification`, jadi heuristik itu selalu `false` dan salah
untuk semua akun. Fix: backend (`get_account_line_history`) sudah lama benar
(pakai `account.normal_balance` asli) — frontend sekarang tinggal memakai
`row.running_balance` yang sudah dihitung server, bukan menebak ulang.
"""

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

from accounting.models import Account, AccountClassification, JournalEntry
from accounting.services.journal import create_journal_entry
from accounting.services.ledger import get_account_line_history
from api.customer_models import Supplier, Customer
from api.models import Contact, Order, OrderActivityLog
from api.pos_models import POSSale
from api.product_models import Purchase, PurchasePayment

User = get_user_model()


class LedgerRunningBalanceNormalBalanceTest(TestCase):
    """Meniru contoh referensi Olsera: 3 baris debit berturut-turut ke 1 akun biaya
    (90.190 lalu 160.000 lalu 9.360.000) harus jadi running balance
    90.190 / 250.190 / 9.610.190 — akumulasi naik, bukan turun."""

    def setUp(self):
        expense_cls = AccountClassification.objects.create(
            name="Harga Pokok Penjualan Test", account_type="expense",
            code_range_start=50000, code_range_end=59999,
        )
        cash_cls = AccountClassification.objects.create(
            name="Kas & Bank Test", account_type="asset",
            code_range_start=10000, code_range_end=19999,
        )
        self.pembelian = Account.objects.create(
            code="51991", name="Pembelian Test", account_type="expense",
            classification=expense_cls,
        )
        self.kas = Account.objects.create(
            code="11992", name="Kas Test", account_type="asset",
            classification=cash_cls,
        )
        self.supplier1 = Supplier.objects.create(nama="Jaya Sentosa, CV")
        self.supplier2 = Supplier.objects.create(nama="Sinar Cemerlang, PT")
        self.customer1 = Customer.objects.create(nama="Budi Pelanggan")

    def _post(self, amount, supplier=None, customer=None):
        create_journal_entry(
            date=date.today(),
            lines=[
                {
                    "account": self.pembelian, "debit": Decimal(amount), "kredit": 0,
                    "description": f"Pembelian dari {supplier.nama if supplier else 'tunai'}",
                    "supplier": supplier, "customer": customer,
                },
                {"account": self.kas, "debit": 0, "kredit": Decimal(amount)},
            ],
            description="TEST-BIAYA",
        )

    def test_running_balance_akumulasi_naik_untuk_akun_debit_normal(self):
        self._post("90190", supplier=self.supplier1)
        self._post("160000", supplier=self.supplier2)
        self._post("9360000", supplier=self.supplier2)

        history = get_account_line_history(self.pembelian, date.today(), date.today())
        rows = history["rows"]
        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["running_balance"], Decimal("90190"))
        self.assertEqual(rows[1]["running_balance"], Decimal("250190"))
        self.assertEqual(rows[2]["running_balance"], Decimal("9610190"))

    def test_pelanggan_supplier_diberi_label_prefix(self):
        self._post("90190", supplier=self.supplier1)
        self._post("50000", customer=self.customer1)

        history = get_account_line_history(self.pembelian, date.today(), date.today())
        rows = history["rows"]
        self.assertEqual(rows[0]["pelanggan_supplier"], "Supplier: Jaya Sentosa, CV")
        self.assertEqual(rows[1]["pelanggan_supplier"], "Pembeli: Budi Pelanggan")

    def test_running_balance_turun_untuk_akun_kredit_normal(self):
        """Kebalikannya: akun kas (debit-normal juga, tapi dites arah lain lewat
        akun kewajiban) — saldo akun kewajiban naik saat KREDIT, bukan debit."""
        liability_cls = AccountClassification.objects.create(
            name="Kewajiban Test", account_type="liability",
            code_range_start=20000, code_range_end=29999,
        )
        hutang = Account.objects.create(
            code="21991", name="Hutang Test", account_type="liability",
            classification=liability_cls,
        )
        create_journal_entry(
            date=date.today(),
            lines=[
                {"account": self.pembelian, "debit": Decimal("100000"), "kredit": 0},
                {"account": hutang, "debit": 0, "kredit": Decimal("100000")},
            ],
            description="TEST-HUTANG-NAIK",
        )
        create_journal_entry(
            date=date.today(),
            lines=[
                {"account": hutang, "debit": Decimal("40000"), "kredit": 0},
                {"account": self.kas, "debit": 0, "kredit": Decimal("40000")},
            ],
            description="TEST-HUTANG-DIBAYAR",
        )
        history = get_account_line_history(hutang, date.today(), date.today())
        rows = history["rows"]
        self.assertEqual(rows[0]["running_balance"], Decimal("100000"))
        self.assertEqual(rows[1]["running_balance"], Decimal("60000"))


class LedgerPelangganSupplierSourceFallbackTest(TestCase):
    """
    Baris jurnal POS/Order/Purchase TIDAK diisi FK customer/supplier eksplisit
    saat posting (Contact/Order tidak kompatibel tipe dengan model Customer,
    lihat _resolve_pelanggan_supplier di services/ledger.py) — tetap harus
    menampilkan nama asli dari transaksi sumber via source_type/source_id,
    bukan kosong. Email hanya terisi kalau kontaknya tertaut ke akun member
    Customer atau supplier punya Supplier master data.
    """

    def setUp(self):
        cash_cls = AccountClassification.objects.create(
            name="Kas Fallback Test", account_type="asset", code_range_start=15000, code_range_end=15999,
        )
        revenue_cls = AccountClassification.objects.create(
            name="Pendapatan Fallback Test", account_type="revenue", code_range_start=45000, code_range_end=45999,
        )
        self.kas = Account.objects.create(code="15001", name="Kas Fallback", account_type="asset", classification=cash_cls)
        self.lawan = Account.objects.create(code="45001", name="Lawan Fallback", account_type="revenue", classification=revenue_cls)

    def _post(self, source_type, source_id):
        create_journal_entry(
            date=date.today(),
            lines=[
                {"account": self.kas, "debit": Decimal("50000"), "kredit": 0},
                {"account": self.lawan, "debit": 0, "kredit": Decimal("50000")},
            ],
            description="Test fallback pelanggan/supplier",
            source_type=source_type,
            source_id=source_id,
        )

    def test_pos_sale_uses_contact_name_and_linked_member_email(self):
        member = Customer.objects.create(nama="Siti Member", email="siti@example.com")
        contact = Contact.objects.create(nomor_wa="628111", nama="Siti Pelanggan", customer=member)
        sale = POSSale.objects.create(nomor="POS-TEST-1", pelanggan=contact)
        self._post(JournalEntry.SourceType.POS_SALE, sale.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["pelanggan_supplier"], "Pembeli: Siti Pelanggan")
        self.assertEqual(row["email"], "siti@example.com")

    def test_order_payment_uses_order_nama_no_email_when_no_linked_contact(self):
        """source_id ORDER_PAYMENT = OrderActivityLog.id, BUKAN Order.id (lihat order_posting.py)."""
        order = Order.objects.create(nomor_wa="628222", nama="Andi Order")
        log = OrderActivityLog.objects.create(order=order, tindakan="PAYMENT", keterangan="Bayar DP")
        self._post(JournalEntry.SourceType.ORDER_PAYMENT, log.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["pelanggan_supplier"], "Pembeli: Andi Order")
        self.assertEqual(row["email"], "")

    def test_purchase_payment_uses_supplier_ref_name_and_email(self):
        """source_id PURCHASE_PAYMENT = PurchasePayment.id, BUKAN Purchase.id (lihat purchase_posting.py)."""
        supplier = Supplier.objects.create(nama="CV Sumber Jaya", email="cv@example.com")
        purchase = Purchase.objects.create(nomor="PUR-TEST-1", tanggal=date.today(), supplier_ref=supplier)
        payment = PurchasePayment.objects.create(purchase=purchase, tanggal=date.today(), nominal=Decimal("50000"))
        self._post(JournalEntry.SourceType.PURCHASE_PAYMENT, payment.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["pelanggan_supplier"], "Supplier: CV Sumber Jaya")
        self.assertEqual(row["email"], "cv@example.com")

    def test_no_source_returns_blank_not_crash(self):
        self._post(JournalEntry.SourceType.MANUAL, None)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["pelanggan_supplier"], "")
        self.assertEqual(row["email"], "")

    def test_pos_sale_shows_dilayani_oleh_when_set(self):
        staff = User.objects.create_user(username="staff_layan_pos", password="pw", role="staff", first_name="Budi")
        sale = POSSale.objects.create(nomor="POS-TEST-LAYAN-1", dilayani_oleh=staff)
        self._post(JournalEntry.SourceType.POS_SALE, sale.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["dilayani_oleh"], "Budi")

    def test_order_payment_shows_dilayani_oleh_when_set(self):
        staff = User.objects.create_user(username="staff_layan_order", password="pw", role="staff", first_name="Wati")
        order = Order.objects.create(nomor_wa="628333", nama="Rudi Order", dilayani_oleh=staff)
        log = OrderActivityLog.objects.create(order=order, tindakan="PAYMENT", keterangan="Bayar DP")
        self._post(JournalEntry.SourceType.ORDER_PAYMENT, log.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["dilayani_oleh"], "Wati")

    def test_dilayani_oleh_blank_when_not_set(self):
        sale = POSSale.objects.create(nomor="POS-TEST-LAYAN-2")
        self._post(JournalEntry.SourceType.POS_SALE, sale.id)

        row = get_account_line_history(self.kas, date.today(), date.today())["rows"][0]
        self.assertEqual(row["dilayani_oleh"], "")


class LedgerAccountExportColumnsTest(TestCase):
    """Export Excel Rincian Mutasi Akun wajib memuat kolom Pelanggan/Supplier
    dan Email (2026-07-31, instruksi user) — sebelumnya kolom ini sengaja
    dihilangkan supaya sama persis PDF Olsera lama."""

    def setUp(self):
        self.owner = User.objects.create_user(username='ledger_exp_owner', password='pw12345', role='owner')
        self.client = APIClient()
        cash_cls = AccountClassification.objects.create(
            name='Kas Export Test', account_type='asset', code_range_start=16000, code_range_end=16999,
        )
        revenue_cls = AccountClassification.objects.create(
            name='Pendapatan Export Test', account_type='revenue', code_range_start=46000, code_range_end=46999,
        )
        self.kas = Account.objects.create(code='16001', name='Kas Export', account_type='asset', classification=cash_cls)
        self.lawan = Account.objects.create(code='46001', name='Lawan Export', account_type='revenue', classification=revenue_cls)

        supplier = Supplier.objects.create(nama='CV Export Jaya', email='export@example.com')
        create_journal_entry(
            date=date.today(),
            lines=[
                {'account': self.kas, 'debit': Decimal('20000'), 'kredit': 0, 'supplier': supplier},
                {'account': self.lawan, 'debit': 0, 'kredit': Decimal('20000')},
            ],
            description='Test export kolom',
        )

    def test_account_export_includes_pelanggan_supplier_and_email_columns(self):
        self.client.force_authenticate(self.owner)
        res = self.client.get(f'/api/accounting/ledger/{self.kas.id}/export/', {
            'date_from': str(date.today()), 'date_to': str(date.today()),
        })
        self.assertEqual(res.status_code, 200)

        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Pelanggan/Supplier', headers)
        self.assertIn('Email', headers)
        self.assertIn('Dilayani Oleh', headers)

        data_row = {headers[i]: cell.value for i, cell in enumerate(ws[2])}
        self.assertEqual(data_row['Pelanggan/Supplier'], 'Supplier: CV Export Jaya')
        self.assertEqual(data_row['Email'], 'export@example.com')
        self.assertFalse(data_row['Dilayani Oleh'])  # openpyxl baca cell string kosong sebagai None
