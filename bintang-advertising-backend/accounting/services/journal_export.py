import io

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("No. Jurnal", 18),
    ("Tanggal", 12),
    ("Kode Akun", 12),
    ("Nama Akun", 30),
    ("Deskripsi", 30),
    ("No. Dokumen", 16),
    ("Debit", 15),
    ("Kredit", 15),
]


def build_journal_export(entries):
    """Bangun workbook xlsx dari daftar JournalEntry (dengan lines-nya) — 1 baris per JournalEntryLine."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jurnal Umum"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = 2
    for entry in entries:
        for line in entry.lines.all():
            ws.cell(row=row_idx, column=1, value=entry.entry_number)
            ws.cell(row=row_idx, column=2, value=entry.date.strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=3, value=line.account.code)
            ws.cell(row=row_idx, column=4, value=line.account.name)
            ws.cell(row=row_idx, column=5, value=line.description or entry.description)
            ws.cell(row=row_idx, column=6, value=line.external_document_no)
            ws.cell(row=row_idx, column=7, value=float(line.debit))
            ws.cell(row=row_idx, column=8, value=float(line.kredit))
            row_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
