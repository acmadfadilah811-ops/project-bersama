import io

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROW_COLUMNS = [("Kode Akun", 14), ("Nama Akun", 32), ("Jumlah", 18)]


def _write_header(ws, columns):
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_section(ws, row_idx, title, rows):
    cell = ws.cell(row=row_idx, column=1, value=title)
    cell.font = Font(bold=True)
    row_idx += 1
    for row in rows:
        ws.cell(row=row_idx, column=1, value=row.get("code", ""))
        ws.cell(row=row_idx, column=2, value=row["name"])
        ws.cell(row=row_idx, column=3, value=float(row["amount"]))
        row_idx += 1
    return row_idx


def _write_total(ws, row_idx, label, amount):
    ws.cell(row=row_idx, column=2, value=label).font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value=float(amount)).font = Font(bold=True)
    return row_idx + 1


def build_balance_sheet_xlsx(data):
    """Neraca — Aset (Lancar/Tidak Lancar), Kewajiban, Modal per section, dengan subtotal & total."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Neraca"
    _write_header(ws, ROW_COLUMNS)

    row_idx = 2
    row_idx = _write_section(ws, row_idx, "ASET LANCAR", data["aset_lancar"])
    row_idx = _write_total(ws, row_idx, "Subtotal Aset Lancar", data["subtotal_aset_lancar"])
    row_idx += 1
    row_idx = _write_section(ws, row_idx, "ASET TIDAK LANCAR", data["aset_tidak_lancar"])
    row_idx = _write_total(ws, row_idx, "Subtotal Aset Tidak Lancar", data["subtotal_aset_tidak_lancar"])
    row_idx = _write_total(ws, row_idx, "TOTAL ASET", data["total_aset"])
    row_idx += 1
    row_idx = _write_section(ws, row_idx, "KEWAJIBAN", data["kewajiban"])
    row_idx = _write_total(ws, row_idx, "Subtotal Kewajiban", data["subtotal_kewajiban"])
    row_idx += 1
    row_idx = _write_section(ws, row_idx, "MODAL", data["modal"])
    row_idx = _write_total(ws, row_idx, "Subtotal Modal", data["subtotal_modal"])
    row_idx = _write_total(ws, row_idx, "TOTAL KEWAJIBAN + MODAL", data["total_kewajiban_modal"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


INCOME_STATEMENT_SECTIONS = [
    ("pendapatan", "PENDAPATAN", "subtotal_pendapatan"),
    ("hpp", "HARGA POKOK PENJUALAN (HPP)", "subtotal_hpp"),
    ("biaya_operasional", "BIAYA OPERASIONAL", "total_biaya_operasional"),
    ("pendapatan_non_operasional", "PENDAPATAN NON-OPERASIONAL", "subtotal_pendapatan_non_operasional"),
    ("biaya_non_operasional", "BIAYA NON-OPERASIONAL", "subtotal_biaya_non_operasional"),
]


def build_income_statement_xlsx(data):
    """Laba Rugi 1 periode — 5 seksi klasifikasi COA dengan subtotal & laba bersih."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laba Rugi"
    _write_header(ws, ROW_COLUMNS)

    row_idx = 2
    for key, title, subtotal_key in INCOME_STATEMENT_SECTIONS:
        row_idx = _write_section(ws, row_idx, title, data[key])
        row_idx = _write_total(ws, row_idx, f"Subtotal {title.title()}", data[subtotal_key])
        row_idx += 1

    row_idx = _write_total(ws, row_idx, "Total Laba Kotor", data["total_laba_kotor"])
    row_idx = _write_total(ws, row_idx, "Total Pendapatan Non-Operasional", data["total_pendapatan_non_operasional"])
    row_idx = _write_total(ws, row_idx, "LABA BERSIH", data["laba_bersih"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
