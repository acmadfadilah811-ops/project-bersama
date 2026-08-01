import io

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

SUMMARY_COLUMNS = [
    ("Nomor Akun", 14), ("Nama Akun", 30), ("Klasifikasi", 20),
    ("Debit", 15), ("Kredit", 15), ("Saldo", 15),
]

# Kolom Pelanggan/Supplier + Email ditambahkan (2026-07-31, instruksi user) —
# PDF asli Olsera (ledger_detail_11101...pdf) tidak punya kolom ini, tapi
# datanya penting untuk rekonsiliasi per pelanggan/supplier. Email kosong
# untuk baris yang sumbernya POS/Order tanpa kontak member ber-email (lihat
# _resolve_pelanggan_supplier di services/ledger.py).
# Kolom Dilayani Oleh ditambahkan (2026-08-01, instruksi user) — nama staf
# yang melayani pelanggan (POSSale.dilayani_oleh/Order.dilayani_oleh), beda
# dari kolom "Transaksi" (siapa yang posting jurnal) — lihat
# _resolve_dilayani_oleh di services/ledger.py.
DETAIL_COLUMNS = [
    ("Tanggal", 12), ("Transaksi", 18), ("Pelanggan/Supplier", 24), ("Email", 24),
    ("Dilayani Oleh", 20),
    ("Deskripsi", 30), ("Debit", 15), ("Kredit", 15), ("Jumlah", 15),
]

ALL_ACCOUNTS_DETAIL_COLUMNS = [
    ("Nomor Akun", 12), ("Nama Akun", 30), ("Tanggal", 12), ("Transaksi", 18),
    ("Pelanggan/Supplier", 24), ("Email", 24), ("Dilayani Oleh", 20),
    ("Deskripsi", 30), ("Debit", 15), ("Kredit", 15), ("Jumlah", 15),
]


def _write_header(ws, columns):
    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def build_ledger_summary_export(accounts, movements):
    """Ringkasan Buku Besar — 1 baris per akun (Debit/Kredit/Saldo dalam periode)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buku Besar"
    _write_header(ws, SUMMARY_COLUMNS)

    row_idx = 2
    for account in accounts:
        movement = movements.get(account.id, {"debit": 0, "kredit": 0, "saldo_akhir": 0})
        ws.cell(row=row_idx, column=1, value=account.code)
        ws.cell(row=row_idx, column=2, value=account.name)
        ws.cell(row=row_idx, column=3, value=account.classification.name if account.classification else "")
        ws.cell(row=row_idx, column=4, value=float(movement["debit"]))
        ws.cell(row=row_idx, column=5, value=float(movement["kredit"]))
        ws.cell(row=row_idx, column=6, value=float(movement["saldo_akhir"]))
        row_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_ledger_all_accounts_detail_export(accounts_with_history):
    """
    Detail Buku Besar — rincian transaksi SEMUA akun, dikelompokkan per akun,
    kronologis per akun (saldo berjalan reset tiap ganti akun). Akun tanpa
    transaksi di rentang tanggal dilewati (tidak ada baris kosong percuma).
    `accounts_with_history`: list of (Account, history_dict dari get_account_line_history).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Buku Besar - Detail"
    _write_header(ws, ALL_ACCOUNTS_DETAIL_COLUMNS)

    row_idx = 2
    for account, history in accounts_with_history:
        if not history["rows"]:
            continue
        for row in history["rows"]:
            ws.cell(row=row_idx, column=1, value=account.code)
            ws.cell(row=row_idx, column=2, value=account.name)
            ws.cell(row=row_idx, column=3, value=row["date"].strftime("%Y-%m-%d"))
            ws.cell(row=row_idx, column=4, value=row["entry_number"])
            ws.cell(row=row_idx, column=5, value=row["pelanggan_supplier"])
            ws.cell(row=row_idx, column=6, value=row["email"])
            ws.cell(row=row_idx, column=7, value=row["dilayani_oleh"])
            ws.cell(row=row_idx, column=8, value=row["description"])
            ws.cell(row=row_idx, column=9, value=float(row["debit"]))
            ws.cell(row=row_idx, column=10, value=float(row["kredit"]))
            ws.cell(row=row_idx, column=11, value=float(row["running_balance"]))
            row_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_ledger_account_export(account, history):
    """Rincian Mutasi Akun — riwayat transaksi 1 akun dengan saldo berjalan."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = account.code[:31]
    _write_header(ws, DETAIL_COLUMNS)

    row_idx = 2
    for row in history["rows"]:
        ws.cell(row=row_idx, column=1, value=row["date"].strftime("%Y-%m-%d"))
        ws.cell(row=row_idx, column=2, value=row["entry_number"])
        ws.cell(row=row_idx, column=3, value=row["pelanggan_supplier"])
        ws.cell(row=row_idx, column=4, value=row["email"])
        ws.cell(row=row_idx, column=5, value=row["dilayani_oleh"])
        ws.cell(row=row_idx, column=6, value=row["description"])
        ws.cell(row=row_idx, column=7, value=float(row["debit"]))
        ws.cell(row=row_idx, column=8, value=float(row["kredit"]))
        ws.cell(row=row_idx, column=9, value=float(row["running_balance"]))
        row_idx += 1

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
