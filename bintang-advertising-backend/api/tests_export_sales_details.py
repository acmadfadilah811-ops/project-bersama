"""
api/tests_export_sales_details.py
Export "Rincian Penjualan" (/api/export/sales-details/) — bug ditemukan &
diperbaiki (2026-08-01, instruksi user): Order.kasir dan Order.diskon tidak
pernah ada di model Order (AttributeError), jadi export ini CRASH 500 setiap
ada Order di rentang tanggal terpilih. Diperbaiki ke field asli
(dilayani_oleh, diskon_persen). Sekalian ditambah kolom Email Pelanggan.
"""
from datetime import date
from io import BytesIO

import openpyxl
from django.contrib.auth import get_user_model
from django.test import TestCase
from rest_framework.test import APIClient

from api.customer_models import Customer
from api.models import Contact, Order, OrderItem
from api.pos_models import POSSale

User = get_user_model()


class ExportSalesDetailsTestCase(TestCase):
    def setUp(self):
        self.owner = User.objects.create_user(username="owner_export_sd", password="pw", role="owner")
        self.client = APIClient()
        self.client.force_authenticate(self.owner)
        self.today = date.today()

    def test_export_with_order_in_range_does_not_crash(self):
        """Bug lama: Order di rentang tanggal bikin export 500 (o.kasir tidak ada)."""
        Order.objects.create(nomor_wa="628999", nama="Pelanggan Order Test")
        OrderItem.objects.create(
            order=Order.objects.first(), jenis_produk="Item Test", harga_jual=100000,
        )

        res = self.client.get(
            "/api/export/sales-details/",
            {"start_date": str(self.today), "end_date": str(self.today)},
        )
        self.assertEqual(res.status_code, 200, res.content)

        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb.active
        self.assertEqual(ws.max_row, 2)  # header + 1 baris Order

    def test_export_includes_email_column_for_pos_and_order(self):
        member = Customer.objects.create(nama="Siti Member", email="siti@example.com")
        contact = Contact.objects.create(nomor_wa="628111", nama="Siti Pelanggan", customer=member)
        POSSale.objects.create(nomor="POS-EXPORT-TEST-1", pelanggan=contact, total=50000)

        member2 = Customer.objects.create(nama="Andi Member", email="andi@example.com")
        Contact.objects.create(nomor_wa="628222", nama="Andi Pelanggan", customer=member2)
        order = Order.objects.create(nomor_wa="628222", nama="Andi Pelanggan")
        OrderItem.objects.create(order=order, jenis_produk="Item Test", harga_jual=75000)

        res = self.client.get(
            "/api/export/sales-details/",
            {"start_date": str(self.today), "end_date": str(self.today)},
        )
        self.assertEqual(res.status_code, 200, res.content)

        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn("Email Pelanggan", headers)
        email_col = headers.index("Email Pelanggan")

        emails = {row[email_col].value for row in ws.iter_rows(min_row=2)}
        self.assertIn("siti@example.com", emails)
        self.assertIn("andi@example.com", emails)

    def test_export_email_blank_when_no_linked_member(self):
        POSSale.objects.create(nomor="POS-EXPORT-TEST-2", total=30000)

        res = self.client.get(
            "/api/export/sales-details/",
            {"start_date": str(self.today), "end_date": str(self.today)},
        )
        self.assertEqual(res.status_code, 200, res.content)

        wb = openpyxl.load_workbook(BytesIO(res.content))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        email_col = headers.index("Email Pelanggan")
        row = list(ws.iter_rows(min_row=2))[0]
        self.assertFalse(row[email_col].value)
